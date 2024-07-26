from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from csvtoxlsx import csv_to_excel

QB_COL = ["A", "B", "C", "D", "E"]
RB_COL = ["G", "H", "I", "J", "K"]
WR_COL = ["M", "N", "O", "P", "Q"]
TE_COL = ["S", "T", "U", "V", "W"]
SOURCE_COL = ["A", "B", "I", "C", "H"]
HEADINGS = ["Name", "Team", "Tier", "Bye", "ADP"]


def move_to_page(pos):
    try:
        wb = load_workbook("D:/Documents/Fantasy Football/2024/All in one.xlsx")
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "data"
        wb.save("D:/Documents/Fantasy Football/2024/All in one.xlsx")

    add_headers(ws, pos)
    try:
        source_wb = load_workbook(
            f"D:/Documents/Fantasy Football/2024/UDK - Position Rankings - {pos}.xlsx"
        )
        source_ws = source_wb.active
    except FileNotFoundError:
        csv_to_excel(pos)
        source_wb = load_workbook(
            f"D:/Documents/Fantasy Football/2024/UDK - Position Rankings - {pos}.xlsx"
        )

    source_ws = source_wb.active

    for row in range(2, source_ws.max_row + 1):
        for x in range(0, 5):
            ws[get_cols(pos)[x] + str(row)] = source_ws[SOURCE_COL[x] + str(row)].value

    wb.save("D:/Documents/Fantasy Football/2024/All in one.xlsx")


def add_headers(ws, pos):
    for x in range(0, 5):
        ws[get_cols(pos)[x] + "1"] = HEADINGS[x]
    return


def get_cols(pos):
    if pos == "QB":
        return QB_COL
    elif pos == "RB":
        return RB_COL
    elif pos == "WR":
        return WR_COL
    elif pos == "TE":
        return TE_COL
    else:
        return None


def make_pretty(pos):
    wb = load_workbook("D:/Documents/Fantasy Football/2024/All in one.xlsx")
    ws = wb.active

    source_wb = load_workbook(
        f"D:/Documents/Fantasy Football/2024/UDK - Position Rankings - {pos}.xlsx"
    )
    source_ws = source_wb.active

    for row in range(2, source_ws.max_row + 1):
        if ws[get_cols(pos)[2] + str(row)].value % 2 == 0:
            for x in range(0, 5):
                ws[get_cols(pos)[x] + str(row)].fill = PatternFill(
                    "solid", fgColor="C6E0B4"
                )
        else:
            for x in range(0, 5):
                ws[get_cols(pos)[x] + str(row)].fill = PatternFill(
                    "solid", fgColor="D9E1F2"
                )

    wb.save("D:/Documents/Fantasy Football/2024/All in one.xlsx")


def main():
    move_to_page("QB")
    move_to_page("RB")
    move_to_page("WR")
    move_to_page("TE")

    make_pretty("QB")
    make_pretty("RB")
    make_pretty("WR")
    make_pretty("TE")


if __name__ == "__main__":
    main()
